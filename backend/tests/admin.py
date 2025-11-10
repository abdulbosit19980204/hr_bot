from django.contrib import admin
from django.utils.html import format_html
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.urls import path
from django import forms
import openpyxl
from openpyxl import Workbook
from .models import Test, Question, AnswerOption, TestResult, UserAnswer


class AnswerOptionInline(admin.TabularInline):
    model = AnswerOption
    extra = 2
    fields = ['text', 'is_correct', 'order']


class TestImportForm(forms.Form):
    excel_file = forms.FileField(label='Excel fayl', help_text='Test, savollar va javoblar bilan Excel fayl')


@admin.register(Test)
class TestAdmin(admin.ModelAdmin):
    list_display = ['title', 'positions_display', 'test_mode', 'time_limit', 'passing_score', 'random_questions_count', 'trial_questions_count', 'is_active', 'questions_count', 'created_at']
    list_filter = ['is_active', 'test_mode', 'positions', 'created_at']
    search_fields = ['title', 'description']
    filter_horizontal = ['positions']
    fieldsets = (
        ('Basic Info', {'fields': ('title', 'description', 'positions')}),
        ('Test Settings', {'fields': ('test_mode', 'time_limit', 'passing_score', 'random_questions_count', 'trial_questions_count', 'show_answers_immediately')}),
        ('Status', {'fields': ('is_active',)}),
        ('Timestamps', {'fields': ('created_at', 'updated_at'), 'classes': ('collapse',)}),
    )
    readonly_fields = ['created_at', 'updated_at']
    
    def positions_display(self, obj):
        return ", ".join([p.name for p in obj.positions.all()[:3]])
    positions_display.short_description = 'Positions'
    change_list_template = 'admin/tests/test_change_list.html'

    def questions_count(self, obj):
        return obj.questions.count()
    questions_count.short_description = 'Questions'

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('import-excel/', self.import_excel, name='tests_test_import_excel'),
            path('export-excel/', self.export_excel, name='tests_test_export_excel'),
        ]
        return custom_urls + urls

    def import_excel(self, request):
        # Example template yuklab olish
        if request.GET.get('download_template') == '1':
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="test_template.xlsx"'
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Test Template"
            
            # Test info headers
            ws.cell(row=1, column=1).value = "Title:"
            ws.cell(row=2, column=1).value = "Description:"
            ws.cell(row=3, column=1).value = "Position:"
            ws.cell(row=4, column=1).value = "Time Limit (minutes):"
            ws.cell(row=5, column=1).value = "Passing Score (%):"
            
            # Example test data
            ws.cell(row=1, column=2).value = "Example Test"
            ws.cell(row=2, column=2).value = "This is an example test description"
            ws.cell(row=3, column=2).value = "Software Engineer"
            ws.cell(row=4, column=2).value = 60
            ws.cell(row=5, column=2).value = 70
            
            # Question headers (row 7)
            headers = ['Question', 'Option 1', 'Option 2', 'Option 3', 'Option 4', 'Correct Answer (1-4)']
            for col, header in enumerate(headers, start=1):
                ws.cell(row=7, column=col).value = header
            
            # Example questions
            example_questions = [
                ["What is Python?", "A programming language", "A snake", "A framework", "A database", 1],
                ["What is Django?", "A web framework", "A database", "A language", "A tool", 1],
                ["What is REST API?", "An API architecture", "A database", "A language", "A framework", 1],
            ]
            
            for row_idx, question_data in enumerate(example_questions, start=8):
                for col_idx, value in enumerate(question_data, start=1):
                    ws.cell(row=row_idx, column=col_idx).value = value
            
            wb.save(response)
            return response
        
        if request.method == 'POST':
            form = TestImportForm(request.POST, request.FILES)
            if form.is_valid():
                excel_file = request.FILES['excel_file']
                try:
                    wb = openpyxl.load_workbook(excel_file)
                    ws = wb.active
                    
                    # First row: Test info (title, description, position, time_limit, passing_score)
                    test_data = {
                        'title': ws.cell(row=1, column=2).value or '',
                        'description': ws.cell(row=2, column=2).value or '',
                        'time_limit': int(ws.cell(row=4, column=2).value or 60),
                        'passing_score': int(ws.cell(row=5, column=2).value or 60),
                    }
                    
                    # Create test
                    test = Test.objects.create(**test_data)
                    
                    # Handle position (positions ManyToMany field)
                    position_name = ws.cell(row=3, column=2).value
                    if position_name:
                        from users.models import Position
                        # Try to get existing position or create new one
                        position, created = Position.objects.get_or_create(
                            name=str(position_name).strip(),
                            defaults={'is_open': True, 'description': ''}
                        )
                        test.positions.add(position)
                    
                    # Questions start from row 8 (row 7 is header)
                    # Format: Question | Option1 | Option2 | Option3 | Option4 | Correct Answer (1-4)
                    current_row = 8
                    question_order = 1
                    
                    # Skip header row (row 7) and start from row 8
                    while current_row <= ws.max_row:
                        question_text = ws.cell(row=current_row, column=1).value
                        if not question_text:
                            current_row += 1
                            continue
                        
                        # Skip if it's a header row (check if it contains "Question" or "Option")
                        if isinstance(question_text, str):
                            question_text_lower = question_text.lower()
                            if 'question' in question_text_lower or 'option' in question_text_lower:
                                current_row += 1
                                continue
                        
                        question = Question.objects.create(
                            test=test,
                            text=str(question_text),
                            order=question_order
                        )
                        
                        # Read options (columns 2-5)
                        options = []
                        correct_index = None
                        for col in range(2, 6):
                            option_text = ws.cell(row=current_row, column=col).value
                            if option_text:
                                options.append(str(option_text))
                        
                        # Get correct answer index (column 6)
                        correct_answer = ws.cell(row=current_row, column=6).value
                        if correct_answer:
                            try:
                                correct_index = int(correct_answer) - 1  # 0-based index
                            except:
                                correct_index = 0
                        
                        # Create answer options
                        for idx, option_text in enumerate(options):
                            AnswerOption.objects.create(
                                question=question,
                                text=option_text,
                                is_correct=(idx == correct_index),
                                order=idx + 1
                            )
                        
                        current_row += 1
                        question_order += 1
                    
                    self.message_user(request, f"Test '{test.title}' muvaffaqiyatli import qilindi!")
                    return redirect('..')
                except Exception as e:
                    self.message_user(request, f"Xatolik: {str(e)}", level='error')
        else:
            form = TestImportForm()
        
        context = {
            'form': form,
            'opts': self.model._meta,
            'has_view_permission': self.has_view_permission(request),
        }
        return render(request, 'admin/tests/import_excel.html', context)

    def export_excel(self, request):
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="test_results.xlsx"'
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Results"
        
        # Headers
        headers = ['ID', 'Foydalanuvchi', 'Test', 'Ball', 'To\'g\'ri javoblar', 
                  'Jami savollar', 'Foiz', 'Holat', 'Sana']
        ws.append(headers)
        
        # Get all results
        results = TestResult.objects.select_related('user', 'test').all()
        
        for result in results:
            row = [
                result.id,
                f"{result.user.first_name} {result.user.last_name}".strip() or result.user.username,
                result.test.title,
                result.score,
                result.correct_answers,
                result.total_questions,
                f"{result.score}%",
                "O'tdi" if result.is_passed else "O'tmadi",
                result.completed_at.strftime('%Y-%m-%d %H:%M:%S')
            ]
            ws.append(row)
        
        wb.save(response)
        return response


@admin.register(Question)
class QuestionAdmin(admin.ModelAdmin):
    list_display = ['text_short', 'test', 'order', 'options_count', 'created_at']
    list_filter = ['test', 'created_at']
    search_fields = ['text', 'test__title']
    inlines = [AnswerOptionInline]
    ordering = ['test', 'order']

    def text_short(self, obj):
        return obj.text[:50] + '...' if len(obj.text) > 50 else obj.text
    text_short.short_description = 'Question'

    def options_count(self, obj):
        return obj.options.count()
    options_count.short_description = 'Options'


@admin.register(AnswerOption)
class AnswerOptionAdmin(admin.ModelAdmin):
    list_display = ['text_short', 'question', 'is_correct', 'order']
    list_filter = ['is_correct', 'question__test']
    search_fields = ['text', 'question__text']
    ordering = ['question', 'order']

    def text_short(self, obj):
        return obj.text[:50] + '...' if len(obj.text) > 50 else obj.text
    text_short.short_description = 'Answer'


class UserAnswerInline(admin.TabularInline):
    model = UserAnswer
    extra = 0
    readonly_fields = ['question', 'selected_option', 'is_correct']
    can_delete = False


@admin.register(TestResult)
class TestResultAdmin(admin.ModelAdmin):
    list_display = ['user', 'test', 'attempt_number', 'score', 'correct_answers', 'total_questions', 'is_passed_display', 'is_completed', 'started_at', 'completed_at']
    list_filter = ['test', 'is_completed', 'attempt_number', 'completed_at', 'user__position']
    search_fields = ['user__username', 'user__first_name', 'user__last_name', 'test__title']
    readonly_fields = ['started_at', 'completed_at', 'time_taken', 'attempt_number']
    inlines = [UserAnswerInline]
    date_hierarchy = 'completed_at'
    actions = ['export_to_excel', 'export_to_csv']

    def is_passed_display(self, obj):
        if obj.is_passed:
            return format_html('<span style="color: green;">✓ Passed</span>')
        return format_html('<span style="color: red;">✗ Failed</span>')
    is_passed_display.short_description = 'Status'

    def get_queryset(self, request):
        return super().get_queryset(request).select_related('user', 'test')

    def export_to_excel(self, request, queryset):
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="test_results.xlsx"'
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Results"
        
        headers = ['ID', 'Foydalanuvchi', 'Email', 'Telefon', 'Lavozim', 'Test', 
                  'Ball', 'To\'g\'ri javoblar', 'Jami savollar', 'Foiz', 'Holat', 'Sana']
        ws.append(headers)
        
        for result in queryset:
            row = [
                result.id,
                f"{result.user.first_name} {result.user.last_name}".strip() or result.user.username,
                result.user.email or '',
                result.user.phone or '',
                result.user.position or '',
                result.test.title,
                result.score,
                result.correct_answers,
                result.total_questions,
                f"{result.score}%",
                "O'tdi" if result.is_passed else "O'tmadi",
                result.completed_at.strftime('%Y-%m-%d %H:%M:%S')
            ]
            ws.append(row)
        
        wb.save(response)
        return response
    export_to_excel.short_description = "Tanlangan natijalarni Excel'ga eksport qilish"

    def export_to_csv(self, request, queryset):
        import csv
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="test_results.csv"'
        response.write('\ufeff')  # BOM for UTF-8
        
        writer = csv.writer(response)
        writer.writerow(['ID', 'Foydalanuvchi', 'Email', 'Telefon', 'Lavozim', 'Test', 
                        'Ball', 'To\'g\'ri javoblar', 'Jami savollar', 'Foiz', 'Holat', 'Sana'])
        
        for result in queryset:
            writer.writerow([
                result.id,
                f"{result.user.first_name} {result.user.last_name}".strip() or result.user.username,
                result.user.email or '',
                result.user.phone or '',
                result.user.position or '',
                result.test.title,
                result.score,
                result.correct_answers,
                result.total_questions,
                f"{result.score}%",
                "O'tdi" if result.is_passed else "O'tmadi",
                result.completed_at.strftime('%Y-%m-%d %H:%M:%S')
            ])
        
        return response
    export_to_csv.short_description = "Tanlangan natijalarni CSV'ga eksport qilish"


@admin.register(UserAnswer)
class UserAnswerAdmin(admin.ModelAdmin):
    list_display = ['result', 'question_short', 'selected_option_short', 'is_correct_display']
    list_filter = ['is_correct', 'result__test']
    search_fields = ['result__user__username', 'question__text']
    readonly_fields = ['result', 'question', 'selected_option', 'is_correct']

    def question_short(self, obj):
        return obj.question.text[:50] + '...' if len(obj.question.text) > 50 else obj.question.text
    question_short.short_description = 'Question'

    def selected_option_short(self, obj):
        return obj.selected_option.text[:50] + '...' if len(obj.selected_option.text) > 50 else obj.selected_option.text
    selected_option_short.short_description = 'Selected Answer'

    def is_correct_display(self, obj):
        if obj.is_correct:
            return format_html('<span style="color: green;">✓</span>')
        return format_html('<span style="color: red;">✗</span>')
    is_correct_display.short_description = 'Correct'

