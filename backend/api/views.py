from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.views import APIView
from django.contrib.auth import get_user_model
from django.db.models import Q, Count, Avg
from django.db import models
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import SearchFilter, OrderingFilter
from rest_framework_simplejwt.tokens import RefreshToken
from rest_framework_simplejwt.views import TokenObtainPairView
from django.utils import timezone
from datetime import timedelta
from django.http import HttpResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import random
import logging
import asyncio

logger = logging.getLogger(__name__)

from users.models import CV, Position, TelegramProfile, Notification
from users.services import send_telegram_message_async, send_notification_to_users
from tests.models import Test, Question, AnswerOption, TestResult
from .serializers import (
    TestSerializer, TestListSerializer, QuestionSerializer,
    UserSerializer, UserCreateSerializer, CVSerializer,
    TestResultSerializer, TestResultCreateSerializer, PositionSerializer,
    NotificationSerializer, NotificationErrorSerializer
)

User = get_user_model()


class PositionViewSet(viewsets.ModelViewSet):
    """Position viewset - superuser uchun to'liq CRUD, boshqalar uchun faqat ochiq positionlar"""
    serializer_class = PositionSerializer
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['is_open']
    search_fields = ['name', 'description']
    ordering_fields = ['name', 'created_at']
    ordering = ['name']
    
    def get_queryset(self):
        """Superuser uchun barcha positionlar, boshqalar uchun faqat ochiq positionlar"""
        if self.request.user.is_authenticated and self.request.user.is_superuser:
            return Position.objects.all()
        return Position.objects.filter(is_open=True)
    
    def get_permissions(self):
        """AllowAny for list/retrieve, IsAuthenticated + is_superuser for create/update/delete"""
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            return [IsAuthenticated()]
        return [AllowAny()]
    
    def create(self, request, *args, **kwargs):
        """Create position - only for superusers"""
        if not request.user.is_superuser:
            return Response(
                {'error': 'Only superusers can create positions'},
                status=status.HTTP_403_FORBIDDEN
            )
        return super().create(request, *args, **kwargs)
    
    def update(self, request, *args, **kwargs):
        """Update position - only for superusers"""
        if not request.user.is_superuser:
            return Response(
                {'error': 'Only superusers can update positions'},
                status=status.HTTP_403_FORBIDDEN
            )
        return super().update(request, *args, **kwargs)
    
    def destroy(self, request, *args, **kwargs):
        """Delete position - only for superusers"""
        if not request.user.is_superuser:
            return Response(
                {'error': 'Only superusers can delete positions'},
                status=status.HTTP_403_FORBIDDEN
            )
        return super().destroy(request, *args, **kwargs)


class TestViewSet(viewsets.ModelViewSet):
    queryset = Test.objects.all().prefetch_related('questions__options', 'positions')
    permission_classes = [AllowAny]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['is_active', 'positions']
    search_fields = ['title', 'description']
    ordering_fields = ['created_at', 'title']
    ordering = ['-created_at']

    def get_serializer_class(self):
        if self.action == 'list':
            return TestListSerializer
        return TestSerializer
    
    def get_permissions(self):
        """
        AllowAny for list/retrieve, IsAuthenticated + is_superuser for create/update/delete
        """
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            from rest_framework.permissions import IsAuthenticated
            return [IsAuthenticated()]
        return [AllowAny()]
    
    def create(self, request, *args, **kwargs):
        """Create test - only for superusers"""
        if not request.user.is_authenticated or not request.user.is_superuser:
            return Response(
                {'error': 'Permission denied. Superuser access required.'},
                status=status.HTTP_403_FORBIDDEN
            )
        return super().create(request, *args, **kwargs)
    
    def update(self, request, *args, **kwargs):
        """Update test - only for superusers"""
        if not request.user.is_authenticated or not request.user.is_superuser:
            return Response(
                {'error': 'Permission denied. Superuser access required.'},
                status=status.HTTP_403_FORBIDDEN
            )
        return super().update(request, *args, **kwargs)
    
    def destroy(self, request, *args, **kwargs):
        """Delete test - only for superusers"""
        if not request.user.is_authenticated or not request.user.is_superuser:
            return Response(
                {'error': 'Permission denied. Superuser access required.'},
                status=status.HTTP_403_FORBIDDEN
            )
        return super().destroy(request, *args, **kwargs)

    def get_queryset(self):
        queryset = super().get_queryset()
        
        # Filter by is_active if provided, otherwise return all tests
        # DjangoFilterBackend will handle is_active filter automatically via filterset_fields
        # But we need to ensure default behavior doesn't filter by is_active
        # The filterset_fields=['is_active'] will handle filtering when is_active param is provided
        
        # Filter by position if provided
        position_id = self.request.query_params.get('position_id')
        if position_id:
            queryset = queryset.filter(positions__id=position_id, positions__is_open=True)
        
        # Filter by test_mode if provided
        test_mode = self.request.query_params.get('test_mode')
        if test_mode:
            queryset = queryset.filter(
                models.Q(test_mode=test_mode) | models.Q(test_mode='both')
            )
        
        # Filter out tests where user has used all attempts
        telegram_id = self.request.query_params.get('telegram_id')
        if telegram_id and self.action == 'list':
            try:
                user = User.objects.get(telegram_id=telegram_id)
                # Get tests where user has completed all attempts
                from tests.models import TestResult
                
                # Get test IDs where user has completed max_attempts (for real tests only)
                completed_tests = []
                for test in queryset:
                    # Check real test attempts (not trial)
                    completed_count = TestResult.objects.filter(
                        user=user,
                        test=test,
                        is_trial=False,
                        is_completed=True
                    ).count()
                    if completed_count >= test.max_attempts:
                        completed_tests.append(test.id)
                
                # Exclude tests where all attempts are used
                if completed_tests:
                    queryset = queryset.exclude(id__in=completed_tests)
            except User.DoesNotExist:
                pass
        
        if self.action == 'retrieve':
            return queryset.prefetch_related('questions__options', 'positions')
        return queryset.prefetch_related('positions')
    
    @action(detail=True, methods=['get'])
    def questions_list(self, request, pk=None):
        """Get all test questions with pagination (for superusers only)"""
        from rest_framework.pagination import PageNumberPagination
        from rest_framework.permissions import IsAdminUser
        
        # Check if user is superuser
        if not request.user.is_authenticated or not request.user.is_superuser:
            return Response(
                {'error': 'Permission denied. Superuser access required.'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        test = self.get_object()
        questions = test.questions.all().prefetch_related('options').order_by('order', 'id')
        
        # Pagination with custom page size
        paginator = PageNumberPagination()
        page_size = request.query_params.get('page_size', '20')
        try:
            page_size = int(page_size)
            if page_size < 1:
                page_size = 20
            if page_size > 100:  # Max limit
                page_size = 100
        except (ValueError, TypeError):
            page_size = 20
        paginator.page_size = page_size
        paginated_questions = paginator.paginate_queryset(questions, request)
        
        serializer = QuestionSerializer(paginated_questions, many=True, context={'admin_view': True, 'request': request})
        return paginator.get_paginated_response(serializer.data)
    
    @action(detail=True, methods=['get'], permission_classes=[IsAuthenticated])
    def export_questions(self, request, pk=None):
        """Export questions to Excel file - only for superusers"""
        if not request.user.is_authenticated or not request.user.is_superuser:
            return Response(
                {'error': 'Permission denied. Superuser access required.'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        test = self.get_object()
        
        try:
            # Get all questions
            questions = test.questions.all().prefetch_related('options').order_by('order', 'id')
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Test Template"
            
            # Test info headers (same format as admin panel)
            ws.cell(row=1, column=1).value = "Title:"
            ws.cell(row=2, column=1).value = "Description:"
            ws.cell(row=3, column=1).value = "Position:"
            ws.cell(row=4, column=1).value = "Time Limit (minutes):"
            ws.cell(row=5, column=1).value = "Passing Score (%):"
            
            # Test data
            ws.cell(row=1, column=2).value = test.title
            ws.cell(row=2, column=2).value = test.description or ""
            # Get first position name
            first_position = test.positions.first()
            ws.cell(row=3, column=2).value = first_position.name if first_position else ""
            ws.cell(row=4, column=2).value = test.time_limit
            ws.cell(row=5, column=2).value = test.passing_score
            
            # Empty row (row 6)
            row = 7
            
            # Question headers (row 7)
            headers = ['Question', 'Option 1', 'Option 2', 'Option 3', 'Option 4', 'Correct Answer (1-4)']
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            row += 1
            
            # Questions (starting from row 8)
            for question in questions:
                options = list(question.options.all().order_by('order', 'id'))
                correct_option_num = None
                
                # Find correct option number (1-4)
                for idx, opt in enumerate(options[:4], start=1):
                    if opt.is_correct:
                        correct_option_num = idx
                        break
                
                # Question text
                ws.cell(row=row, column=1).value = question.text
                
                # Options (max 4)
                for idx, opt in enumerate(options[:4], start=1):
                    ws.cell(row=row, column=idx + 1).value = opt.text
                
                # Correct answer (1-4)
                ws.cell(row=row, column=6).value = correct_option_num if correct_option_num else ""
                
                row += 1
            
            # Auto-adjust column widths
            for col in range(1, 7):
                ws.column_dimensions[chr(64 + col)].width = 30
            
            # Create response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            filename = f"test_{test.id}_questions_{timezone.now().strftime('%Y%m%d')}.xlsx"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            wb.save(response)
            return response
            
        except Exception as e:
            logger.error(f'Error exporting questions: {str(e)}')
            return Response(
                {'error': f'Export failed: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['post'], permission_classes=[IsAuthenticated])
    def import_questions(self, request, pk=None):
        """Import questions from Excel file to existing test - only for superusers"""
        if not request.user.is_authenticated or not request.user.is_superuser:
            return Response(
                {'error': 'Permission denied. Superuser access required.'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        test = self.get_object()
        
        try:
            if 'file' not in request.FILES:
                return Response(
                    {'error': 'Excel file is required'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            excel_file = request.FILES['file']
            
            # Load workbook
            wb = load_workbook(excel_file, data_only=True)
            ws = wb.active
            
            imported_count = 0
            errors = []
            
            # Find questions start row (row 7 is header, row 8+ are questions)
            start_row = 8
            # Check if row 7 has headers
            header_cell = ws.cell(row=7, column=1).value
            if header_cell and ('Question' in str(header_cell) or 'Savol' in str(header_cell)):
                start_row = 8
            else:
                # Try to find header row
                for row in range(1, 20):
                    cell_value = ws.cell(row=row, column=1).value
                    if cell_value and ('Question' in str(cell_value) or 'Savol' in str(cell_value)):
                        start_row = row + 1
                        break
            
            # Process questions
            for row in range(start_row, ws.max_row + 1):
                question_text = ws.cell(row=row, column=1).value
                
                # Skip empty rows
                if not question_text or not str(question_text).strip():
                    continue
                
                # Skip if it's a header row
                if isinstance(question_text, str):
                    question_text_lower = str(question_text).lower()
                    if 'question' in question_text_lower or 'option' in question_text_lower or 'savol' in question_text_lower:
                        continue
                
                try:
                    # Get options
                    options = []
                    for col in range(2, 6):  # Columns B-E (Option 1-4)
                        option_text = ws.cell(row=row, column=col).value
                        if option_text and str(option_text).strip():
                            options.append(str(option_text).strip())
                    
                    if not options:
                        errors.append(f"Qator {row}: Variantlar topilmadi")
                        continue
                    
                    # Get correct answer (column 6)
                    correct_answer = ws.cell(row=row, column=6).value
                    correct_option_num = None
                    if correct_answer:
                        try:
                            correct_option_num = int(correct_answer)
                        except (ValueError, TypeError):
                            pass
                    
                    if not correct_option_num or correct_option_num < 1 or correct_option_num > len(options):
                        errors.append(f"Qator {row}: To'g'ri javob noto'g'ri (1-{len(options)} orasida bo'lishi kerak)")
                        continue
                    
                    # Create question
                    question = Question.objects.create(
                        test=test,
                        text=str(question_text).strip(),
                        order=imported_count
                    )
                    
                    # Create answer options
                    for idx, opt_text in enumerate(options, start=1):
                        AnswerOption.objects.create(
                            question=question,
                            text=opt_text,
                            is_correct=(idx == correct_option_num),
                            order=idx - 1
                        )
                    
                    imported_count += 1
                    
                except Exception as e:
                    errors.append(f"Qator {row}: {str(e)}")
            
            response_data = {
                'success': True,
                'imported_count': imported_count,
                'test_id': test.id,
                'test_title': test.title
            }
            
            if errors:
                response_data['errors'] = errors
            
            return Response(response_data, status=status.HTTP_201_CREATED)
            
        except Exception as e:
            logger.error(f'Error importing questions: {str(e)}')
            return Response(
                {'error': f'Import failed: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @action(detail=False, methods=['post'], permission_classes=[IsAuthenticated])
    def import_test(self, request):
        """Import full test from Excel file - only for superusers"""
        if not request.user.is_authenticated or not request.user.is_superuser:
            return Response(
                {'error': 'Permission denied. Superuser access required.'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        try:
            if 'file' not in request.FILES:
                return Response(
                    {'error': 'Excel file is required'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            excel_file = request.FILES['file']
            
            # Load workbook
            wb = load_workbook(excel_file, data_only=True)
            ws = wb.active
            
            # Read test info (rows 1-5)
            test_title = ws.cell(row=1, column=2).value or ''
            test_description = ws.cell(row=2, column=2).value or ''
            position_name = ws.cell(row=3, column=2).value or ''
            time_limit = ws.cell(row=4, column=2).value
            passing_score = ws.cell(row=5, column=2).value
            
            # Create new test
            try:
                time_limit = int(time_limit) if time_limit else 60
                passing_score = int(passing_score) if passing_score else 60
            except (ValueError, TypeError):
                time_limit = 60
                passing_score = 60
            
            test_data = {
                'title': str(test_title).strip() if test_title else 'Imported Test',
                'description': str(test_description).strip() if test_description else '',
                'time_limit': time_limit,
                'passing_score': passing_score,
            }
            
            test = Test.objects.create(**test_data)
            
            # Handle position
            if position_name:
                from users.models import Position
                position, created = Position.objects.get_or_create(
                    name=str(position_name).strip(),
                    defaults={'is_open': True, 'description': ''}
                )
                test.positions.add(position)
            
            imported_count = 0
            errors = []
            
            # Find questions start row (row 7 is header, row 8+ are questions)
            start_row = 8
            # Check if row 7 has headers
            header_cell = ws.cell(row=7, column=1).value
            if header_cell and ('Question' in str(header_cell) or 'Savol' in str(header_cell)):
                start_row = 8
            else:
                # Try to find header row
                for row in range(1, 20):
                    cell_value = ws.cell(row=row, column=1).value
                    if cell_value and ('Question' in str(cell_value) or 'Savol' in str(cell_value)):
                        start_row = row + 1
                        break
            
            # Process questions
            for row in range(start_row, ws.max_row + 1):
                question_text = ws.cell(row=row, column=1).value
                
                # Skip empty rows
                if not question_text or not str(question_text).strip():
                    continue
                
                # Skip if it's a header row
                if isinstance(question_text, str):
                    question_text_lower = str(question_text).lower()
                    if 'question' in question_text_lower or 'option' in question_text_lower or 'savol' in question_text_lower:
                        continue
                
                try:
                    # Get options
                    options = []
                    for col in range(2, 6):  # Columns B-E (Option 1-4)
                        option_text = ws.cell(row=row, column=col).value
                        if option_text and str(option_text).strip():
                            options.append(str(option_text).strip())
                    
                    if not options:
                        errors.append(f"Qator {row}: Variantlar topilmadi")
                        continue
                    
                    # Get correct answer (column 6)
                    correct_answer = ws.cell(row=row, column=6).value
                    correct_option_num = None
                    if correct_answer:
                        try:
                            correct_option_num = int(correct_answer)
                        except (ValueError, TypeError):
                            pass
                    
                    if not correct_option_num or correct_option_num < 1 or correct_option_num > len(options):
                        errors.append(f"Qator {row}: To'g'ri javob noto'g'ri (1-{len(options)} orasida bo'lishi kerak)")
                        continue
                    
                    # Create question
                    question = Question.objects.create(
                        test=test,
                        text=str(question_text).strip(),
                        order=imported_count
                    )
                    
                    # Create answer options
                    for idx, opt_text in enumerate(options, start=1):
                        AnswerOption.objects.create(
                            question=question,
                            text=opt_text,
                            is_correct=(idx == correct_option_num),
                            order=idx - 1
                        )
                    
                    imported_count += 1
                    
                except Exception as e:
                    errors.append(f"Qator {row}: {str(e)}")
            
            response_data = {
                'success': True,
                'imported_count': imported_count,
                'test_id': test.id,
                'test_title': test.title,
                'test_created': True
            }
            
            if errors:
                response_data['errors'] = errors
            
            return Response(response_data, status=status.HTTP_201_CREATED)
            
        except Exception as e:
            logger.error(f'Error importing test: {str(e)}')
            return Response(
                {'error': f'Import failed: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @action(detail=True, methods=['get'])
    def questions(self, request, pk=None):
        """Get test questions (random if configured)"""
        test = self.get_object()
        is_trial = request.query_params.get('trial', 'false').lower() == 'true'
        telegram_id = request.query_params.get('telegram_id')
        
        # Check if user is blocked
        if telegram_id:
            try:
                user = User.objects.get(telegram_id=telegram_id)
                if user.is_blocked:
                    return Response(
                        {'error': 'User is blocked', 'reason': user.blocked_reason},
                        status=status.HTTP_403_FORBIDDEN
                    )
                
                # Check trial test attempts
                if is_trial:
                    # Check trial attempts count (main check based on completed results)
                    trial_results = TestResult.objects.filter(
                        user=user,
                        test=test,
                        is_trial=True,
                        is_completed=True
                    ).count()
                    
                    if trial_results >= test.max_trial_attempts:
                        return Response(
                            {
                                'error': 'All trial attempts used',
                                'message': f'Siz trial testni {trial_results} marta ishlagansiz. Ruxsat etilgan urinishlar soni: {test.max_trial_attempts}',
                                'attempts_used': trial_results,
                                'max_trial_attempts': test.max_trial_attempts
                            },
                            status=status.HTTP_400_BAD_REQUEST
                        )
                else:
                    # Check real test attempts
                    completed_results = TestResult.objects.filter(
                        user=user,
                        test=test,
                        is_trial=False,
                        is_completed=True
                    ).count()
                    
                    if completed_results >= test.max_attempts:
                        return Response(
                            {
                                'error': 'All attempts used',
                                'message': f'Siz bu testni {completed_results} marta ishlagansiz. Ruxsat etilgan urinishlar soni: {test.max_attempts}',
                                'attempts_used': completed_results,
                                'max_attempts': test.max_attempts
                            },
                            status=status.HTTP_400_BAD_REQUEST
                        )
            except User.DoesNotExist:
                pass
        
        questions = test.questions.all().prefetch_related('options').order_by('order', 'id')
        
        # Trial test
        if is_trial:
            questions = list(questions)
            trial_count = test.trial_questions_count
            if len(questions) > trial_count:
                questions = random.sample(questions, trial_count)
        # Random questions for regular test
        elif test.random_questions_count > 0:
            questions = list(questions)
            if len(questions) > test.random_questions_count:
                questions = random.sample(questions, test.random_questions_count)
        
        # Shuffle questions order randomly
        questions = list(questions)
        random.shuffle(questions)
        
        serializer = QuestionSerializer(questions, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def start_test(self, request, pk=None):
        """Start test session or resume existing test"""
        from django.utils import timezone
        from datetime import timedelta
        
        test = self.get_object()
        telegram_id = request.data.get('telegram_id')
        is_trial = request.data.get('trial', False)
        
        if telegram_id:
            try:
                user = User.objects.get(telegram_id=telegram_id)
                if user.is_blocked:
                    return Response(
                        {'error': 'User is blocked', 'reason': user.blocked_reason},
                        status=status.HTTP_403_FORBIDDEN
                    )
                
                # Check trial test attempts
                if is_trial:
                    # Check trial attempts count (main check based on completed results)
                    trial_results = TestResult.objects.filter(
                        user=user,
                        test=test,
                        is_trial=True,
                        is_completed=True
                    ).count()
                    
                    if trial_results >= test.max_trial_attempts:
                        return Response(
                            {
                                'error': 'All trial attempts used',
                                'message': f'Siz trial testni {trial_results} marta ishlagansiz. Ruxsat etilgan urinishlar soni: {test.max_trial_attempts}',
                                'attempts_used': trial_results,
                                'max_trial_attempts': test.max_trial_attempts
                            },
                            status=status.HTTP_400_BAD_REQUEST
                        )
                else:
                    # Check real test attempts
                    completed_results = TestResult.objects.filter(
                        user=user,
                        test=test,
                        is_trial=False,
                        is_completed=True
                    ).count()
                    
                    if completed_results >= test.max_attempts:
                        return Response(
                            {
                                'error': 'All attempts used',
                                'message': f'Siz bu testni {completed_results} marta ishlagansiz. Ruxsat etilgan urinishlar soni: {test.max_attempts}',
                                'attempts_used': completed_results,
                                'max_attempts': test.max_attempts
                            },
                            status=status.HTTP_400_BAD_REQUEST
                        )
                
                # Check if user has uploaded CV for this test (if passed)
                # If CV is uploaded, don't allow retaking the test
                from users.models import CV
                from tests.models import TestResult
                passed_results = TestResult.objects.filter(
                    user=user,
                    test=test,
                    is_completed=True,
                    score__gte=test.passing_score
                ).exists()
                
                if passed_results:
                    # Check if CV is uploaded
                    has_cv = CV.objects.filter(user=user).exists()
                    if has_cv:
                        return Response(
                            {'error': 'CV already uploaded. Test cannot be retaken.'},
                            status=status.HTTP_400_BAD_REQUEST
                        )
                
                # Check for incomplete test (resume)
                incomplete_result = TestResult.objects.filter(
                    user=user,
                    test=test,
                    is_completed=False
                ).order_by('-started_at').first()
                
                if incomplete_result:
                    # Check if test time limit has expired
                    time_elapsed = (timezone.now() - incomplete_result.started_at).total_seconds() / 60
                    if time_elapsed < test.time_limit:
                        # Resume existing test
                        return Response({
                            'session_token': str(incomplete_result.id),  # Use result ID as session token
                            'test_id': test.id,
                            'time_limit': test.time_limit,
                            'time_elapsed': int(time_elapsed),
                            'time_remaining': int(test.time_limit - time_elapsed),
                            'is_trial': is_trial,
                            'resume': True,
                            'result_id': incomplete_result.id,
                            'attempt_number': incomplete_result.attempt_number
                        })
                    else:
                        # Time expired, mark as completed
                        incomplete_result.is_completed = True
                        incomplete_result.completed_at = timezone.now()
                        incomplete_result.save()
                
                # Start new test attempt
                attempt_number = TestResult.objects.filter(
                    user=user,
                    test=test
                ).count() + 1
                
                # Create new test result (incomplete)
                from tests.models import TestResult
                new_result = TestResult.objects.create(
                    user=user,
                    test=test,
                    score=0,
                    total_questions=0,
                    correct_answers=0,
                    time_taken=0,
                    attempt_number=attempt_number,
                    is_completed=False,
                    is_trial=is_trial,
                    started_at=timezone.now()
                )
                
                # Generate test session token
                import uuid
                session_token = str(uuid.uuid4())
                
                return Response({
                    'session_token': session_token,
                    'test_id': test.id,
                    'time_limit': test.time_limit,
                    'is_trial': is_trial,
                    'resume': False,
                    'result_id': new_result.id,
                    'attempt_number': attempt_number
                })
            except User.DoesNotExist:
                return Response(
                    {'error': 'User not found'},
                    status=status.HTTP_404_NOT_FOUND
                )
        
        return Response(
            {'error': 'telegram_id required'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    @action(detail=True, methods=['post'])
    def notify_page_leave(self, request, pk=None):
        """Notify about page leave attempt during test"""
        test = self.get_object()
        telegram_id = request.data.get('telegram_id')
        attempts = request.data.get('attempts', 1)
        test_id = request.data.get('test_id')
        
        if not telegram_id:
            return Response(
                {'error': 'telegram_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            user = User.objects.get(telegram_id=telegram_id)
            
            # Log page leave attempt
            test_title = test.title if test else 'Unknown'
            
            # Send notification to admin (log as ERROR for visibility)
            if attempts >= 2:
                logger.error(
                    f"🚨 CRITICAL: User attempted to leave test page {attempts} times - "
                    f"User: {user.first_name} {user.last_name} (ID: {user.id}, Telegram ID: {telegram_id}), "
                    f"Test: {test_title} (ID: {test_id or pk})"
                )
            else:
                logger.warning(
                    f"⚠️ Page leave attempt: User {user.first_name} {user.last_name} "
                    f"(Telegram ID: {telegram_id}) attempted to leave test page {attempts} time(s) - "
                    f"Test: {test_title} (ID: {test_id or pk})"
                )
            
            # If 2+ attempts, block user immediately
            if attempts >= 2:
                user.is_blocked = True
                user.blocked_reason = f"Test tark etildi (cheating) - {attempts} marta urinish"
                user.blocked_at = timezone.now()
                user.save()
                
                logger.error(
                    f"🚨 User blocked due to page leave: user_id={user.id}, "
                    f"telegram_id={telegram_id}, test_id={test_id or pk}, attempts={attempts}, "
                    f"reason: {user.blocked_reason}"
                )
                
                # Try to send notification to Telegram bot admin group
                # This is done via logging - Telegram bot can monitor logs or we can add webhook
                # For now, we'll log it and the bot's admin notification system will handle it
                
                return Response({
                    'message': 'User blocked due to multiple page leave attempts',
                    'blocked': True,
                    'reason': user.blocked_reason
                }, status=status.HTTP_200_OK)
            
            return Response({
                'message': 'Page leave attempt logged',
                'attempts': attempts,
                'blocked': False
            }, status=status.HTTP_200_OK)
            
        except User.DoesNotExist:
            return Response(
                {'error': 'User not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            logger.error(f"Error notifying page leave: {e}", exc_info=True)
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['post'])
    def block_user(self, request, pk=None):
        """Block user for cheating"""
        test = self.get_object()
        telegram_id = request.data.get('telegram_id')
        reason = request.data.get('reason', 'Test tark etildi (cheating)')
        
        if telegram_id:
            try:
                user = User.objects.get(telegram_id=telegram_id)
                user.is_blocked = True
                user.blocked_reason = reason
                user.blocked_at = timezone.now()
                user.save()
                
                return Response({
                    'message': 'User blocked successfully',
                    'user_id': user.id
                })
            except User.DoesNotExist:
                return Response(
                    {'error': 'User not found'},
                    status=status.HTTP_404_NOT_FOUND
                )
        
            return Response(
                {'error': 'telegram_id required'},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @action(detail=False, methods=['get'], permission_classes=[IsAuthenticated])
    def export_excel(self, request):
        """Export tests to Excel - only for authenticated users"""
        from django.http import HttpResponse
        from openpyxl import Workbook
        from django.utils import timezone
        
        # Get filtered queryset
        queryset = self.filter_queryset(self.get_queryset())
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"tests_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Tests"
        
        headers = ['ID', 'Test nomi', 'Tavsif', 'Lavozimlar', 'Savollar soni', 'Vaqt chegarasi (daqiqa)', 
                  'O\'tish balli (%)', 'Max urinishlar', 'Test rejimi', 'Status', 'Yaratilgan sana']
        ws.append(headers)
        
        for test in queryset:
            positions = ', '.join([p.name for p in test.positions.all()]) if test.positions.exists() else ''
            test_mode = {
                'telegram': 'Telegram',
                'web': 'Web',
                'both': 'Ikkalasi'
            }.get(test.test_mode, test.test_mode)
            
            row = [
                test.id,
                test.title or '',
                test.description or '',
                positions,
                test.questions.count() if hasattr(test, 'questions') else 0,
                test.time_limit or 0,
                test.passing_score or 0,
                test.max_attempts or 0,
                test_mode,
                'Aktiv' if test.is_active else 'Nofaol',
                test.created_at.strftime('%Y-%m-%d %H:%M:%S') if test.created_at else ''
            ]
            ws.append(row)
        
        wb.save(response)
        return response
    
    @action(detail=False, methods=['get'], permission_classes=[IsAuthenticated])
    def export_csv(self, request):
        """Export tests to CSV - only for authenticated users"""
        import csv
        from django.http import HttpResponse
        from django.utils import timezone
        
        # Get filtered queryset
        queryset = self.filter_queryset(self.get_queryset())
        
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        filename = f"tests_{timezone.now().strftime('%Y%m%d_%H%M%S')}.csv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response.write('\ufeff')  # BOM for UTF-8
        
        writer = csv.writer(response)
        writer.writerow(['ID', 'Test nomi', 'Tavsif', 'Lavozimlar', 'Savollar soni', 'Vaqt chegarasi (daqiqa)', 
                        'O\'tish balli (%)', 'Max urinishlar', 'Test rejimi', 'Status', 'Yaratilgan sana'])
        
        for test in queryset:
            positions = ', '.join([p.name for p in test.positions.all()]) if test.positions.exists() else ''
            test_mode = {
                'telegram': 'Telegram',
                'web': 'Web',
                'both': 'Ikkalasi'
            }.get(test.test_mode, test.test_mode)
            
            writer.writerow([
                test.id,
                test.title or '',
                test.description or '',
                positions,
                test.questions.count() if hasattr(test, 'questions') else 0,
                test.time_limit or 0,
                test.passing_score or 0,
                test.max_attempts or 0,
                test_mode,
                'Aktiv' if test.is_active else 'Nofaol',
                test.created_at.strftime('%Y-%m-%d %H:%M:%S') if test.created_at else ''
            ])
        
        return response


class UserViewSet(viewsets.ModelViewSet):
    queryset = User.objects.all()
    permission_classes = [AllowAny]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['position']
    search_fields = ['username', 'first_name', 'last_name', 'email', 'phone']
    ordering_fields = ['created_at']
    ordering = ['-created_at']

    def get_serializer_class(self):
        if self.action == 'create':
            return UserCreateSerializer
        return UserSerializer

    @action(detail=False, methods=['get'], permission_classes=[IsAuthenticated])
    def me(self, request):
        """Get current authenticated user"""
        serializer = UserSerializer(request.user)
        return Response(serializer.data)
    
    @action(detail=False, methods=['post'])
    def telegram_auth(self, request):
        """Authenticate user by Telegram ID and update Telegram info"""
        telegram_id = request.data.get('telegram_id')
        if not telegram_id:
            return Response({'error': 'telegram_id is required'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            user = User.objects.get(telegram_id=telegram_id)
            # Update TelegramProfile, not User's first_name/last_name
            try:
                telegram_profile = TelegramProfile.objects.get(telegram_id=telegram_id)
                # Update existing profile
                telegram_profile.user = user
            except TelegramProfile.DoesNotExist:
                # Create new profile
                telegram_profile = TelegramProfile.objects.create(
                    user=user,
                    telegram_id=telegram_id,
                    telegram_is_premium=False,
                    telegram_is_bot=False
                )
            
            # Update TelegramProfile info if provided
            if 'telegram_username' in request.data:
                telegram_profile.telegram_username = request.data.get('telegram_username')
            if 'telegram_language_code' in request.data:
                telegram_profile.telegram_language_code = request.data.get('telegram_language_code')
            
            # Always ensure telegram_is_premium is set (NOT NULL constraint)
            # Set it explicitly to avoid None values
            if 'telegram_is_premium' in request.data:
                telegram_profile.telegram_is_premium = bool(request.data.get('telegram_is_premium', False))
            else:
                # If not provided, ensure it's set to False (not None)
                telegram_profile.telegram_is_premium = False if telegram_profile.telegram_is_premium is None else bool(telegram_profile.telegram_is_premium)
            
            # Always ensure telegram_is_bot is set (NOT NULL constraint)
            telegram_profile.telegram_is_bot = False if telegram_profile.telegram_is_bot is None else bool(telegram_profile.telegram_is_bot)
            
            # Update telegram first_name and last_name (NOT User's first_name/last_name)
            if 'first_name' in request.data:
                telegram_first_name = request.data.get('first_name', '')
                telegram_profile.telegram_first_name = telegram_first_name if telegram_first_name is not None else ''
            if 'last_name' in request.data:
                telegram_last_name = request.data.get('last_name', '')
                telegram_profile.telegram_last_name = telegram_last_name if telegram_last_name is not None else ''
            
            # Update user's telegram_id if not set
            if not user.telegram_id:
                user.telegram_id = telegram_id
                user.save()
            
            telegram_profile.save()
            
            refresh = RefreshToken.for_user(user)
            return Response({
                'refresh': str(refresh),
                'access': str(refresh.access_token),
                'user': UserSerializer(user).data
            })
        except User.DoesNotExist:
            # Create user if not exists (without telegram first_name/last_name)
            user = User.objects.create_user(
                username=f'user_{telegram_id}',
                telegram_id=telegram_id,
                # Don't set first_name/last_name from telegram - user will fill it during registration
                first_name='',
                last_name=''
            )
            
            # Create TelegramProfile with telegram info
            telegram_first_name = request.data.get('first_name', '') or ''
            telegram_last_name = request.data.get('last_name', '') or ''
            if telegram_first_name is None:
                telegram_first_name = ''
            if telegram_last_name is None:
                telegram_last_name = ''
            
            telegram_profile = TelegramProfile.objects.create(
                user=user,
                telegram_id=telegram_id,
                telegram_first_name=telegram_first_name,
                telegram_last_name=telegram_last_name,
                telegram_username=request.data.get('telegram_username'),
                telegram_language_code=request.data.get('telegram_language_code'),
                telegram_is_premium=request.data.get('telegram_is_premium', False),
                telegram_is_bot=False
            )
            
            refresh = RefreshToken.for_user(user)
            return Response({
                'refresh': str(refresh),
                'access': str(refresh.access_token),
                'user': UserSerializer(user).data
            }, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['post'])
    def create_telegram_user(self, request):
        """Create user from Telegram data"""
        telegram_id = request.data.get('telegram_id')
        if telegram_id:
            try:
                user = User.objects.get(telegram_id=telegram_id)
                # Update user data
                for key in ['first_name', 'last_name', 'email', 'phone']:
                    if key in request.data:
                        setattr(user, key, request.data[key])
                user.save()
                return Response(UserSerializer(user).data)
            except User.DoesNotExist:
                user = User.objects.create_user(
                    username=f'user_{telegram_id}',
                    telegram_id=telegram_id,
                    first_name=request.data.get('first_name', ''),
                    last_name=request.data.get('last_name', ''),
                    email=request.data.get('email', ''),
                    phone=request.data.get('phone', '')
                )
                return Response(UserSerializer(user).data, status=status.HTTP_201_CREATED)
        return Response({'error': 'telegram_id required'}, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['get'], permission_classes=[IsAuthenticated])
    def export_excel(self, request):
        """Export users to Excel - only for authenticated users"""
        from django.http import HttpResponse
        from openpyxl import Workbook
        from django.utils import timezone
        
        # Get filtered queryset
        queryset = self.filter_queryset(self.get_queryset())
        
        # If not staff, only allow export of own data
        if not request.user.is_staff:
            queryset = queryset.filter(id=request.user.id)
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"users_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Users"
        
        headers = ['ID', 'Username', 'Ism', 'Familiya', 'Email', 'Telefon', 'Telegram ID', 'Lavozim', 'Status', 'Yaratilgan sana']
        ws.append(headers)
        
        for user in queryset:
            position_name = ''
            if user.position:
                position_name = str(user.position.name) if hasattr(user.position, 'name') else str(user.position)
            
            row = [
                user.id,
                user.username or '',
                user.first_name or '',
                user.last_name or '',
                user.email or '',
                user.phone or '',
                user.telegram_id or '',
                position_name,
                'Aktiv' if user.is_active else 'Nofaol',
                user.date_joined.strftime('%Y-%m-%d %H:%M:%S') if user.date_joined else ''
            ]
            ws.append(row)
        
        wb.save(response)
        return response
    
    @action(detail=False, methods=['get'], permission_classes=[IsAuthenticated])
    def export_csv(self, request):
        """Export users to CSV - only for authenticated users"""
        import csv
        from django.http import HttpResponse
        from django.utils import timezone
        
        # Get filtered queryset
        queryset = self.filter_queryset(self.get_queryset())
        
        # If not staff, only allow export of own data
        if not request.user.is_staff:
            queryset = queryset.filter(id=request.user.id)
        
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        filename = f"users_{timezone.now().strftime('%Y%m%d_%H%M%S')}.csv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response.write('\ufeff')  # BOM for UTF-8
        
        writer = csv.writer(response)
        writer.writerow(['ID', 'Username', 'Ism', 'Familiya', 'Email', 'Telefon', 'Telegram ID', 'Lavozim', 'Status', 'Yaratilgan sana'])
        
        for user in queryset:
            position_name = ''
            if user.position:
                position_name = str(user.position.name) if hasattr(user.position, 'name') else str(user.position)
            
            writer.writerow([
                user.id,
                user.username or '',
                user.first_name or '',
                user.last_name or '',
                user.email or '',
                user.phone or '',
                user.telegram_id or '',
                position_name,
                'Aktiv' if user.is_active else 'Nofaol',
                user.date_joined.strftime('%Y-%m-%d %H:%M:%S') if user.date_joined else ''
            ])
        
        return response


class QuestionViewSet(viewsets.ModelViewSet):
    queryset = Question.objects.all().prefetch_related('options')
    serializer_class = QuestionSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter]
    filterset_fields = ['test']
    search_fields = ['text']
    
    def get_serializer_context(self):
        """Add admin_view context for serializer"""
        context = super().get_serializer_context()
        context['admin_view'] = True
        return context
    
    def get_permissions(self):
        """
        AllowAny for list/retrieve, IsAuthenticated + is_superuser for create/update/delete
        """
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            return [IsAuthenticated()]
        return [IsAuthenticated()]
    
    def perform_create(self, serializer):
        """Set test from request data"""
        test_id = self.request.data.get('test')
        if test_id:
            from tests.models import Test
            try:
                test = Test.objects.get(pk=test_id)
                serializer.save(test=test)
            except Test.DoesNotExist:
                from rest_framework.exceptions import ValidationError
                raise ValidationError({'test': 'Test topilmadi'})
        else:
            serializer.save()
    
    def create(self, request, *args, **kwargs):
        """Create question - only for superusers"""
        if not request.user.is_superuser:
            return Response(
                {'error': 'Permission denied. Only superusers can create questions.'},
                status=status.HTTP_403_FORBIDDEN
            )
        return super().create(request, *args, **kwargs)
    
    def update(self, request, *args, **kwargs):
        """Update question - only for superusers"""
        if not request.user.is_superuser:
            return Response(
                {'error': 'Permission denied. Only superusers can update questions.'},
                status=status.HTTP_403_FORBIDDEN
            )
        return super().update(request, *args, **kwargs)
    
    def partial_update(self, request, *args, **kwargs):
        """Partial update question - only for superusers"""
        if not request.user.is_superuser:
            return Response(
                {'error': 'Permission denied. Only superusers can update questions.'},
                status=status.HTTP_403_FORBIDDEN
            )
        return super().partial_update(request, *args, **kwargs)
    
    def destroy(self, request, *args, **kwargs):
        """Delete question - only for superusers"""
        if not request.user.is_superuser:
            return Response(
                {'error': 'Permission denied. Only superusers can delete questions.'},
                status=status.HTTP_403_FORBIDDEN
            )
        return super().destroy(request, *args, **kwargs)


class CVViewSet(viewsets.ModelViewSet):
    serializer_class = CVSerializer
    permission_classes = [AllowAny]  # Bot uchun ochiq qildik
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    filterset_fields = ['user']
    ordering_fields = ['uploaded_at']
    ordering = ['-uploaded_at']

    def get_queryset(self):
        # Staff uchun barcha CV'lar
        if self.request.user.is_authenticated and self.request.user.is_staff:
            return CV.objects.all()
        
        # Telegram ID bo'yicha filter (bot uchun)
        telegram_id = self.request.query_params.get('user__telegram_id')
        if telegram_id:
            return CV.objects.filter(user__telegram_id=telegram_id)
        
        # Authenticated user uchun faqat o'z CV'lari
        if self.request.user.is_authenticated:
            return CV.objects.filter(user=self.request.user)
        
        return CV.objects.none()

    def perform_create(self, serializer):
        # Bot uchun telegram_id orqali user topish
        telegram_id = self.request.data.get('telegram_id')
        if telegram_id:
            try:
                user = User.objects.get(telegram_id=telegram_id)
                serializer.save(user=user)
            except User.DoesNotExist:
                from rest_framework import serializers as drf_serializers
                raise drf_serializers.ValidationError("User not found")
        else:
            if not self.request.user.is_authenticated:
                from rest_framework import serializers as drf_serializers
                raise drf_serializers.ValidationError("Authentication required")
            serializer.save(user=self.request.user)
    
    def create(self, request, *args, **kwargs):
        response = super().create(request, *args, **kwargs)
        
        # CV yuklanganidan keyin xabar
        if response.status_code == status.HTTP_201_CREATED:
            response.data['message'] = (
                "✅ CV muvaffaqiyatli yuklandi!\n\n"
                "Biz sizga tez orada aloqaga chiqamiz va siz bilan birinchi Zoom interview uchun maslahatlarimizni beramiz.\n\n"
                "Rahmat!"
            )
        
        return response
    
    @action(detail=False, methods=['get'], permission_classes=[IsAuthenticated])
    def export_excel(self, request):
        """Export CVs to Excel - only for authenticated users"""
        from django.http import HttpResponse
        from openpyxl import Workbook
        from django.utils import timezone
        
        # Get filtered queryset
        queryset = self.filter_queryset(self.get_queryset())
        
        # If not staff, only allow export of own CVs
        if not request.user.is_staff:
            queryset = queryset.filter(user=request.user)
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"cvs_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb = Workbook()
        ws = wb.active
        ws.title = "CVs"
        
        headers = ['ID', 'Foydalanuvchi', 'Email', 'Telefon', 'Fayl nomi', 'Fayl hajmi (KB)', 'Yuklangan sana']
        ws.append(headers)
        
        for cv in queryset:
            user_name = f"{cv.user.first_name} {cv.user.last_name}".strip() or cv.user.username if cv.user else ''
            
            # Get file size in KB
            file_size_kb = 0
            if cv.file:
                try:
                    file_size_kb = round(cv.file.size / 1024, 2) if hasattr(cv.file, 'size') else 0
                except:
                    file_size_kb = 0
            
            row = [
                cv.id,
                user_name,
                cv.user.email if cv.user else '',
                cv.user.phone if cv.user else '',
                cv.file.name.split('/')[-1] if cv.file else '',
                file_size_kb,
                cv.uploaded_at.strftime('%Y-%m-%d %H:%M:%S') if cv.uploaded_at else ''
            ]
            ws.append(row)
        
        wb.save(response)
        return response
    
    @action(detail=False, methods=['post'], permission_classes=[IsAuthenticated])
    def download_zip(self, request):
        """Download CV files as ZIP - only for authenticated users"""
        import zipfile
        import io
        import os
        from django.http import HttpResponse
        from django.utils import timezone
        from django.core.files.storage import default_storage
        
        # Get CV IDs from request (can be list or single ID)
        cv_ids = request.data.get('cv_ids', [])
        if not cv_ids:
            from rest_framework import status
            from rest_framework.response import Response
            return Response(
                {'error': 'CV ID\'lar ko\'rsatilmagan'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # If single ID, convert to list
        if not isinstance(cv_ids, list):
            cv_ids = [cv_ids]
        
        # Get CVs
        queryset = CV.objects.filter(id__in=cv_ids)
        
        # If not staff, only allow download of own CVs
        if not request.user.is_staff:
            queryset = queryset.filter(user=request.user)
        
        if not queryset.exists():
            from rest_framework import status
            from rest_framework.response import Response
            return Response(
                {'error': 'CV\'lar topilmadi yoki ruxsat yo\'q'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Create ZIP file in memory
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for cv in queryset:
                if cv.file and default_storage.exists(cv.file.name):
                    try:
                        # Get file content
                        file_content = default_storage.open(cv.file.name).read()
                        
                        # Create filename: UserName_CV_ID_OriginalFileName
                        user_name = f"{cv.user.first_name}_{cv.user.last_name}".strip() if cv.user else ''
                        if not user_name and cv.user:
                            user_name = cv.user.username
                        if not user_name:
                            user_name = 'Unknown'
                        
                        # Clean filename (remove invalid characters)
                        user_name = "".join(c for c in user_name if c.isalnum() or c in (' ', '-', '_')).strip()
                        user_name = user_name.replace(' ', '_')
                        
                        original_filename = cv.file.name.split('/')[-1]
                        zip_filename = f"{user_name}_CV_{cv.id}_{original_filename}"
                        
                        # Add file to ZIP
                        zip_file.writestr(zip_filename, file_content)
                    except Exception as e:
                        # Skip files that can't be read
                        continue
        
        # Prepare response
        zip_buffer.seek(0)
        response = HttpResponse(zip_buffer.read(), content_type='application/zip')
        filename = f"cvs_{timezone.now().strftime('%Y%m%d_%H%M%S')}.zip"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response


class TestResultViewSet(viewsets.ModelViewSet):
    permission_classes = [AllowAny]  # Bot uchun ochiq qildik, filter orqali cheklaymiz
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['test', 'user']
    search_fields = ['user__username', 'user__first_name', 'user__last_name', 'test__title']
    ordering_fields = ['completed_at', 'score']
    ordering = ['-completed_at']

    def get_serializer_class(self):
        if self.action == 'create':
            return TestResultCreateSerializer
        return TestResultSerializer

    def get_queryset(self):
        # Staff uchun barcha natijalar
        if self.request.user.is_authenticated and self.request.user.is_staff:
            return TestResult.objects.filter(is_completed=True).select_related('user', 'test')
        
        # Telegram ID bo'yicha filter (bot uchun)
        telegram_id = self.request.query_params.get('user__telegram_id')
        if telegram_id:
            return TestResult.objects.filter(
                user__telegram_id=telegram_id,
                is_completed=True
            ).select_related('user', 'test')
        
        # Authenticated user uchun faqat o'z natijalari
        if self.request.user.is_authenticated:
            return TestResult.objects.filter(
                user=self.request.user,
                is_completed=True
            ).select_related('user', 'test')
        
        # Unauthenticated - bo'sh queryset
        return TestResult.objects.none()

    def create(self, request, *args, **kwargs):
        # Bot uchun telegram_id qo'shish (request.data dan olish)
        serializer = self.get_serializer(data=request.data, context={'request': request})
        serializer.is_valid(raise_exception=True)
        result = serializer.save()
        
        response_data = TestResultSerializer(result).data
        
        # CV upload request (minimal ball o'tganda)
        if result.is_passed:
            response_data['requires_cv'] = True
            response_data['cv_upload_message'] = (
                "Tabriklaymiz! Siz testdan o'tdingiz. "
                "CV yuklash uchun tayyor bo'ling."
            )
        
        return Response(response_data, status=status.HTTP_201_CREATED)
    
    @action(detail=False, methods=['get'], permission_classes=[IsAuthenticated])
    def export_excel(self, request):
        """Export test results to Excel - only for authenticated users"""
        from django.http import HttpResponse
        from openpyxl import Workbook
        from django.utils import timezone
        
        # Get filtered queryset
        queryset = self.filter_queryset(self.get_queryset())
        
        # If not staff, only allow export of own results
        if not request.user.is_staff:
            queryset = queryset.filter(user=request.user)
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"test_results_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Results"
        
        headers = ['ID', 'Foydalanuvchi', 'Email', 'Telefon', 'Lavozim', 'Test', 
                  'Ball', 'To\'g\'ri javoblar', 'Jami savollar', 'Foiz', 'Holat', 'Sana']
        ws.append(headers)
        
        for result in queryset:
            # Handle position - convert to string safely
            position_name = ''
            if result.user.position:
                position_name = str(result.user.position.name) if hasattr(result.user.position, 'name') else str(result.user.position)
            
            row = [
                result.id,
                f"{result.user.first_name} {result.user.last_name}".strip() or result.user.username,
                result.user.email or '',
                result.user.phone or '',
                position_name,
                result.test.title,
                result.score,
                result.correct_answers,
                result.total_questions,
                f"{result.score}%",
                "O'tdi" if result.is_passed else "O'tmadi",
                result.completed_at.strftime('%Y-%m-%d %H:%M:%S') if result.completed_at else ''
            ]
            ws.append(row)
        
        wb.save(response)
        return response
    
    @action(detail=False, methods=['get'], permission_classes=[IsAuthenticated])
    def export_csv(self, request):
        """Export test results to CSV - only for authenticated users"""
        import csv
        from django.http import HttpResponse
        from django.utils import timezone
        
        # Get filtered queryset
        queryset = self.filter_queryset(self.get_queryset())
        
        # If not staff, only allow export of own results
        if not request.user.is_staff:
            queryset = queryset.filter(user=request.user)
        
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        filename = f"test_results_{timezone.now().strftime('%Y%m%d_%H%M%S')}.csv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response.write('\ufeff')  # BOM for UTF-8
        
        writer = csv.writer(response)
        writer.writerow(['ID', 'Foydalanuvchi', 'Email', 'Telefon', 'Lavozim', 'Test', 
                        'Ball', 'To\'g\'ri javoblar', 'Jami savollar', 'Foiz', 'Holat', 'Sana'])
        
        for result in queryset:
            # Handle position - convert to string safely
            position_name = ''
            if result.user.position:
                position_name = str(result.user.position.name) if hasattr(result.user.position, 'name') else str(result.user.position)
            
            writer.writerow([
                result.id,
                f"{result.user.first_name} {result.user.last_name}".strip() or result.user.username,
                result.user.email or '',
                result.user.phone or '',
                position_name,
                result.test.title,
                result.score,
                result.correct_answers,
                result.total_questions,
                f"{result.score}%",
                "O'tdi" if result.is_passed else "O'tmadi",
                result.completed_at.strftime('%Y-%m-%d %H:%M:%S') if result.completed_at else ''
            ])
        
        return response


class StatisticsView(APIView):
    """Statistics view - Production-safe aggregated statistics"""
    permission_classes = [IsAuthenticated]  # Production'da authentication talab qilamiz
    authentication_classes = []  # Allow token authentication

    def get(self, request):
        today = timezone.now().date()
        now = timezone.now()
        week_ago = now - timedelta(days=7)
        month_ago = now - timedelta(days=30)
        
        # A. Real-time stats
        tests_today = TestResult.objects.filter(completed_at__date=today).count()
        new_users_today = User.objects.filter(date_joined__date=today).count()
        cv_uploads_today = CV.objects.filter(uploaded_at__date=today).count()
        active_users = User.objects.filter(is_active=True, is_blocked=False).count()
        blocked_users = User.objects.filter(is_blocked=True).count()
        
        # B. Trend Charts - Daily test completions (7 days)
        daily_tests = []
        for i in range(6, -1, -1):
            date = today - timedelta(days=i)
            count = TestResult.objects.filter(completed_at__date=date).count()
            daily_tests.append({'date': date.strftime('%Y-%m-%d'), 'count': count})
        
        # User growth (daily, weekly, monthly)
        user_growth = {
            'today': new_users_today,
            'this_week': User.objects.filter(date_joined__gte=week_ago).count(),
            'this_month': User.objects.filter(date_joined__gte=month_ago).count()
        }
        
        # CV upload trends
        cv_trends = {
            'today': cv_uploads_today,
            'this_week': CV.objects.filter(uploaded_at__gte=week_ago).count(),
            'this_month': CV.objects.filter(uploaded_at__gte=month_ago).count()
        }
        
        # C. Comparison Stats
        # Trial vs Real test counts and average scores
        trial_tests = TestResult.objects.filter(test__is_trial=True)
        real_tests = TestResult.objects.filter(test__is_trial=False)
        trial_count = trial_tests.count()
        real_count = real_tests.count()
        trial_avg_score = trial_tests.aggregate(avg=Avg('score'))['avg'] or 0
        real_avg_score = real_tests.aggregate(avg=Avg('score'))['avg'] or 0
        
        # Tests by position (aggregated, no sensitive data)
        tests_by_position = TestResult.objects.values('test__positions__name').annotate(
            count=Count('id'),
            avg_score=Avg('score')
        ).order_by('-count')[:10]
        
        # Test pass rates
        total_results = TestResult.objects.count()
        passed_results = TestResult.objects.filter(score__gte=models.F('test__passing_score')).count()
        pass_rate = (passed_results / total_results * 100) if total_results > 0 else 0
        
        # Top 10 hardest/easiest tests (by average score)
        test_stats = TestResult.objects.values('test__id', 'test__title').annotate(
            avg_score=Avg('score'),
            count=Count('id')
        ).filter(count__gte=5).order_by('avg_score')[:10]  # Hardest
        hardest_tests = list(test_stats)
        
        test_stats_easy = TestResult.objects.values('test__id', 'test__title').annotate(
            avg_score=Avg('score'),
            count=Count('id')
        ).filter(count__gte=5).order_by('-avg_score')[:10]  # Easiest
        easiest_tests = list(test_stats_easy)
        
        # D. Performance Metrics
        avg_time_per_test = TestResult.objects.aggregate(
            avg=Avg('time_taken')
        )['avg'] or 0
        
        fastest_completion = TestResult.objects.order_by('time_taken').first()
        slowest_completion = TestResult.objects.order_by('-time_taken').first()
        
        # Average attempts per user
        attempts_per_user = TestResult.objects.values('user__id').annotate(
            count=Count('id')
        ).aggregate(avg=Avg('count'))['avg'] or 0
        
        users_with_multiple_attempts = User.objects.annotate(
            attempt_count=Count('testresult')
        ).filter(attempt_count__gt=1).count()
        
        # E. User Engagement
        trial_participation = trial_count
        real_participation = real_count
        
        # F. CV Statistics
        total_cvs = CV.objects.count()
        cvs_this_week = CV.objects.filter(uploaded_at__gte=week_ago).count()
        avg_file_size = CV.objects.aggregate(avg=Avg('file_size'))['avg'] or 0
        users_with_cv = User.objects.filter(cv__isnull=False).distinct().count()
        users_passed_and_cv = User.objects.filter(
            testresult__score__gte=models.F('testresult__test__passing_score'),
            cv__isnull=False
        ).distinct().count()
        
        # G. Notifications
        total_notifications = Notification.objects.count()
        sent_notifications = Notification.objects.exclude(sent_at__isnull=True).count()
        draft_notifications = Notification.objects.filter(sent_at__isnull=True).count()
        total_successful_sends = Notification.objects.aggregate(
            total=Count('id'),
            successful=Count('successful_sends')
        )
        success_rate = (total_successful_sends['successful'] / total_successful_sends['total'] * 100) if total_successful_sends['total'] > 0 else 0
        
        # Top notification errors (by type) - aggregated
        from users.models import NotificationError
        top_errors = NotificationError.objects.values('error_type').annotate(
            count=Count('id')
        ).order_by('-count')[:5]
        
        # Overall stats
        total_tests = TestResult.objects.count()
        avg_score = TestResult.objects.aggregate(avg=Avg('score'))['avg'] or 0
        total_users = User.objects.count()
        
        # Best results (top 10, aggregated data only)
        best_results = TestResult.objects.select_related('user', 'test').order_by('-score')[:10]
        best_results_data = [{
            'id': r.id,
            'user_name': f"{r.user.first_name} {r.user.last_name}" if r.user else 'Noma\'lum',
            'test_title': r.test.title if r.test else 'Noma\'lum',
            'score': r.score,
            'correct_answers': r.correct_answers,
            'total_questions': r.total_questions,
            'completed_at': r.completed_at.isoformat() if r.completed_at else None
        } for r in best_results]

        return Response({
            # A. Real-time stats
            'tests_today': tests_today,
            'new_users_today': new_users_today,
            'cv_uploads_today': cv_uploads_today,
            'active_users': active_users,
            'blocked_users': blocked_users,
            
            # B. Trend Charts
            'daily_test_completions': daily_tests,
            'user_growth': user_growth,
            'cv_upload_trends': cv_trends,
            
            # C. Comparison Stats
            'trial_vs_real': {
                'trial_count': trial_count,
                'real_count': real_count,
                'trial_avg_score': round(trial_avg_score, 2),
                'real_avg_score': round(real_avg_score, 2)
            },
            'tests_by_position': list(tests_by_position),
            'pass_rate': round(pass_rate, 2),
            'hardest_tests': hardest_tests,
            'easiest_tests': easiest_tests,
            
            # D. Performance Metrics
            'avg_time_per_test': round(avg_time_per_test, 2) if avg_time_per_test else 0,
            'fastest_completion': fastest_completion.time_taken if fastest_completion else None,
            'slowest_completion': slowest_completion.time_taken if slowest_completion else None,
            'avg_attempts_per_user': round(attempts_per_user, 2),
            'users_with_multiple_attempts': users_with_multiple_attempts,
            
            # E. User Engagement
            'trial_participation': trial_participation,
            'real_participation': real_participation,
            
            # F. CV Statistics
            'total_cvs': total_cvs,
            'cvs_this_week': cvs_this_week,
            'avg_file_size': round(avg_file_size / 1024, 2) if avg_file_size else 0,  # KB
            'users_with_cv': users_with_cv,
            'users_passed_and_cv': users_passed_and_cv,
            
            # G. Notifications
            'total_notifications': total_notifications,
            'sent_notifications': sent_notifications,
            'draft_notifications': draft_notifications,
            'total_successful_sends': total_successful_sends.get('successful', 0),
            'success_rate': round(success_rate, 2),
            'top_notification_errors': list(top_errors),
            
            # Overall stats (backward compatibility)
            'total_tests': total_tests,
            'avg_score': round(avg_score, 2),
            'total_users': total_users,
            'tests_this_week': TestResult.objects.filter(completed_at__gte=week_ago).count(),
            'tests_by_position_old': list(TestResult.objects.values('user__position__name').annotate(
                count=Count('id')
            ).order_by('-count')[:10]),
            'best_results': best_results_data,
        })


class NotificationViewSet(viewsets.ReadOnlyModelViewSet):
    """Notification ViewSet - list, retrieve, filter, search"""
    queryset = Notification.objects.all().prefetch_related('recipients', 'errors', 'created_by').order_by('-created_at')
    serializer_class = NotificationSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['send_to_all', 'created_by']
    search_fields = ['title', 'message']
    ordering_fields = ['created_at', 'sent_at', 'title']
    ordering = ['-created_at']


class NotificationView(APIView):
    """Send notification to selected users"""
    permission_classes = [IsAuthenticated]  # Faqat authenticated userlar
    
    def post(self, request):
        """
        Send notification to selected users
        Expected data:
        {
            "user_ids": [1, 2, 3],
            "title": "Suxbat taklifi",
            "message": "Sizni suxbatga taklif qilamiz...",
            "notification_type": "interview" or "job_offer"
        }
        """
        user_ids = request.data.get('user_ids', [])
        title = request.data.get('title', '')
        message = request.data.get('message', '')
        notification_type = request.data.get('notification_type', 'interview')  # 'interview' or 'job_offer'
        
        if not user_ids:
            return Response(
                {'error': 'user_ids is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if not title or not message:
            return Response(
                {'error': 'title and message are required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # Get users
            users = User.objects.filter(id__in=user_ids, telegram_id__isnull=False)
            if not users.exists():
                return Response(
                    {'error': 'No users found with telegram_id'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Format message based on notification type
            formatted_message = message
            if notification_type == 'interview':
                # Suxbat taklifi uchun format
                formatted_message = f"🎯 <b>Suxbat taklifi</b>\n\n{message}"
            elif notification_type == 'job_offer':
                # Ishga taklif uchun format
                formatted_message = f"💼 <b>Ishga taklif</b>\n\n{message}"
            elif notification_type == 'encouragement':
                # Tashakkur va rag'batlantirish uchun format
                formatted_message = f"🙏 <b>Tashakkur</b>\n\n{message}"
            
            # Create notification
            notification = Notification.objects.create(
                title=title,
                message=formatted_message,
                send_to_all=False,
                created_by=request.user if request.user.is_authenticated else None
            )
            
            # Add recipients
            notification.recipients.set(users)
            
            # Send notifications
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            result = loop.run_until_complete(send_notification_to_users(notification))
            loop.close()
            
            # Update notification stats
            notification.sent_at = timezone.now()
            notification.total_recipients = result['total']
            notification.successful_sends = result['successful']
            notification.failed_sends = result['failed']
            notification.save()
            
            return Response({
                'success': True,
                'message': 'Notification sent successfully',
                'total': result['total'],
                'successful': result['successful'],
                'failed': result['failed'],
                'notification_id': notification.id
            })
            
        except Exception as e:
            logger.error(f"Error sending notification: {e}", exc_info=True)
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
